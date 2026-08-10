#!/usr/bin/env python3
"""
Importa catálogos desde archivos Excel hacia SQL Server.

Cada catálogo se carga en su tabla de staging y luego un procedimiento hace el UPSERT,
igual que el ETL de tickets: no se borra nada, lo que deja de venir en el Excel se marca
con VigenteEnOrigen = 0.

Requisitos previos:
    1. Ejecutar 06_catalogos_excel.sql en SQL Server
    2. pip install openpyxl

Uso:
    python importar_catalogos_excel.py --config config.json
    python importar_catalogos_excel.py --config config.json --catalogo lideres
    python importar_catalogos_excel.py --revisar          # solo lee los Excel, no toca la BD
    python importar_catalogos_excel.py --config config.json --carpeta "C:\\catalogos"
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
import uuid
from pathlib import Path

RAIZ = Path(__file__).resolve().parent
LOG = logging.getLogger("catalogos")

# Definición de cada catálogo: de dónde sale y a dónde va.
# Para agregar otro Excel, basta con añadir una entrada aquí.
CATALOGOS = {
    "grupos": {
        "archivo": "Cat_gruposvalidos.xlsx",
        "hoja": "tabla valid",
        "encabezados": ["Grupo Correcto", "Grupo Valido"],   # esperados en la fila 1
        "columnas": ["GrupoCorrecto", "GrupoValido"],        # nombres en SQL
        "tabla_staging": "stg.CatGruposValidos",
        "sp": "dbo.usp_CargarCatGruposValidos",
        "obligatorias": ["GrupoCorrecto", "GrupoValido"],    # sin estas, la fila se descarta
        "max_len": 150,
    },
    "lideres": {
        "archivo": "lider_grupo.xlsx",
        "hoja": "Torres y lideres",
        "encabezados": ["grupo", "lider"],
        "columnas": ["Grupo", "Lider"],
        "tabla_staging": "stg.CatLiderGrupo",
        "sp": "dbo.usp_CargarCatLiderGrupo",
        "obligatorias": ["Grupo"],
        "max_len": 150,
    },
}


def configurar_log(nivel="INFO"):
    logging.basicConfig(
        level=getattr(logging, nivel.upper(), logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(RAIZ / "logs" / "catalogos.log", encoding="utf-8")
                  if (RAIZ / "logs").exists() else logging.NullHandler(),
                  logging.StreamHandler(sys.stdout)])


def leer_excel(ruta: Path, hoja: str, encabezados: list[str],
               columnas: list[str], obligatorias: list[str], max_len: int) -> list[tuple]:
    """Lee la hoja y devuelve filas limpias, validando que los encabezados coincidan."""
    import openpyxl

    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró {ruta}")

    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    if hoja not in wb.sheetnames:
        raise RuntimeError(f"{ruta.name}: no existe la hoja '{hoja}'. "
                           f"Hojas disponibles: {', '.join(wb.sheetnames)}")
    ws = wb[hoja]
    filas_crudas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not filas_crudas:
        raise RuntimeError(f"{ruta.name}: la hoja '{hoja}' está vacía.")

    # Validar encabezados: si el Excel cambia de estructura, mejor fallar aquí
    cab = [str(c).strip().lower() if c is not None else "" for c in filas_crudas[0][:len(encabezados)]]
    esperados = [e.strip().lower() for e in encabezados]
    if cab != esperados:
        raise RuntimeError(
            f"{ruta.name}: los encabezados no son los esperados.\n"
            f"  esperado: {encabezados}\n  encontrado: {filas_crudas[0][:len(encabezados)]}\n"
            f"  Si el archivo cambió a propósito, ajusta 'encabezados' en CATALOGOS.")

    limpias, descartadas, truncadas = [], 0, 0
    vistas: set[tuple] = set()
    for cruda in filas_crudas[1:]:
        vals = []
        for i in range(len(columnas)):
            v = cruda[i] if i < len(cruda) else None
            v = None if v is None else str(v).strip()
            v = v or None
            if v and len(v) > max_len:
                v = v[:max_len]; truncadas += 1
            vals.append(v)

        if all(v is None for v in vals):
            continue                                   # fila en blanco al final de la hoja
        faltan = [c for c, v in zip(columnas, vals) if c in obligatorias and not v]
        if faltan:
            descartadas += 1
            LOG.warning("Fila descartada, falta %s: %s", ", ".join(faltan), vals)
            continue

        t = tuple(vals)
        if t in vistas:                                 # duplicado exacto en el Excel
            continue
        vistas.add(t)
        limpias.append(t)

    LOG.info("%s · hoja '%s': %s filas útiles (%s descartadas, %s duplicadas, %s truncadas)",
             ruta.name, hoja, len(limpias), descartadas,
             len(filas_crudas) - 1 - len(limpias) - descartadas, truncadas)
    return limpias


def cargar(cn, cat: dict, filas: list[tuple], lote: uuid.UUID) -> tuple[int, int]:
    cols = cat["columnas"] + ["LoteCarga"]
    marcas = ", ".join(["?"] * len(cols))
    cur = cn.cursor()
    cur.execute(f"DELETE FROM {cat['tabla_staging']};")
    cur.fast_executemany = True
    cur.executemany(
        f"INSERT INTO {cat['tabla_staging']} ({', '.join(cols)}) VALUES ({marcas})",
        [f + (str(lote),) for f in filas])
    cn.commit()

    cur.execute("{CALL " + cat["sp"] + " (?)}", str(lote))
    fila = cur.fetchone()
    while fila is None and cur.nextset():
        fila = cur.fetchone()
    cn.commit()
    cur.close()
    return (int(fila[0]), int(fila[1])) if fila else (0, 0)


def main() -> int:
    ap = argparse.ArgumentParser(description="Importa catálogos de Excel a SQL Server")
    ap.add_argument("--config", default=str(RAIZ / "config.json"))
    ap.add_argument("--carpeta", default=str(RAIZ), help="Dónde están los .xlsx")
    ap.add_argument("--catalogo", choices=list(CATALOGOS),
                    help="Cargar sólo uno (por defecto, todos)")
    ap.add_argument("--revisar", action="store_true",
                    help="Sólo lee y valida los Excel; no se conecta a la base")
    ap.add_argument("--nivel-log", default="INFO")
    args = ap.parse_args()

    configurar_log(args.nivel_log)
    carpeta = Path(args.carpeta)
    elegidos = {args.catalogo: CATALOGOS[args.catalogo]} if args.catalogo else CATALOGOS

    # 1. Leer y validar todos los Excel ANTES de tocar la base:
    #    si uno viene mal, no se carga nada a medias.
    datos: dict[str, list[tuple]] = {}
    for nombre, cat in elegidos.items():
        try:
            datos[nombre] = leer_excel(carpeta / cat["archivo"], cat["hoja"],
                                       cat["encabezados"], cat["columnas"],
                                       cat["obligatorias"], cat["max_len"])
        except Exception as e:
            LOG.error("[%s] %s", nombre, e)
            return 1

    if args.revisar:
        print("\nRevisión (no se tocó la base):")
        for nombre, filas in datos.items():
            print(f"\n  {nombre}: {len(filas)} filas · primeras 5")
            for f in filas[:5]:
                print("   ", f)
        return 0

    # 2. Cargar
    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    sys.path.insert(0, str(RAIZ))
    import etl_proactivanet as m

    cn = m.conectar(cfg["sql"])
    rc = 0
    try:
        for nombre, cat in elegidos.items():
            lote = uuid.uuid4()
            try:
                ins, upd = cargar(cn, cat, datos[nombre], lote)
                LOG.info("[%s] listo -> %s nuevos, %s actualizados (de %s filas)",
                         nombre, ins, upd, len(datos[nombre]))
            except Exception as e:
                cn.rollback()
                LOG.exception("[%s] falló la carga: %s", nombre, e)
                rc = 1
    finally:
        cn.close()

    if rc == 0:
        LOG.info("Catálogos importados correctamente.")
    return rc


if __name__ == "__main__":
    sys.exit(main())
